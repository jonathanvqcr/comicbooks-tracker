#!/usr/bin/env python3
"""Parse comic_collection_tracker.xlsx and generate src/data/collection.json."""

import json
import os
import re
from collections import defaultdict

import openpyxl

XLSX_PATH = os.path.expanduser("~/Downloads/comic_collection_tracker.xlsx")
EBAY_FILES = [
    os.path.expanduser(f"~/Downloads/Ebay_Purchase_History_{y}.xlsx")
    for y in [2024, 2025, 2026]
]
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "src", "data", "collection.json")


def slugify(name: str) -> str:
    s = name.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)

    # --- Parse "All Purchases" sheet ---
    ws = wb["All Purchases"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data = rows[1:]

    cover_a: dict[str, set[int]] = defaultdict(set)
    all_volumes: dict[str, set[int]] = defaultdict(set)

    for row in data:
        series = row[1]
        vol = row[2]
        cover = row[3]
        if series is None or vol is None:
            continue
        vol = int(vol)
        all_volumes[series].add(vol)
        if cover and "CVR A" in str(cover).upper():
            cover_a[series].add(vol)

    # --- Parse "Collection Summary" for total issues range ---
    ws2 = wb["Collection Summary"]
    summary_rows = list(ws2.iter_rows(values_only=True))
    summary_header = summary_rows[0]
    summary_data = summary_rows[1:]

    # Build max volume per series from summary "Volumes Owned" column
    summary_max: dict[str, int] = {}
    for row in summary_data:
        series_name = row[0]
        if series_name is None:
            continue
        # Volumes Owned is like "1-17" or "1, 15" or "1"
        vols_str = str(row[1])
        nums = [int(x) for x in re.findall(r"\d+", vols_str)]
        if nums:
            summary_max[series_name] = max(nums)

    wb.close()

    # --- Build item_name -> (series, vol, cover) mapping for image lookup ---
    item_to_info: dict[str, tuple[str, int, str]] = {}
    for row in data:
        if row[1] and row[2] and row[4]:
            item_to_info[str(row[4]).strip()] = (row[1], int(row[2]), row[3] or "")

    # --- Extract image URLs from eBay files ---
    # Per-series: one image (prefer Cover A)
    # Per-issue: series -> vol -> {cover, url} (prefer Cover A)
    series_images: dict[str, str] = {}
    issue_images: dict[str, dict[int, str]] = defaultdict(dict)  # series -> vol -> best url

    for ebay_file in EBAY_FILES:
        if not os.path.exists(ebay_file):
            continue
        ewb = openpyxl.load_workbook(ebay_file, read_only=True)
        ews = ewb[ewb.sheetnames[0]]
        for row in list(ews.iter_rows(values_only=True))[1:]:
            item_name = (row[4] or "").strip()
            img_url = row[11]
            if not img_url or not item_name:
                continue
            info = item_to_info.get(item_name)
            if not info:
                continue
            series_name, vol, cover = info
            is_cover_a = "CVR A" in cover.upper()

            # Series-level image
            if series_name not in series_images or is_cover_a:
                series_images[series_name] = img_url

            # Issue-level image (prefer Cover A)
            if vol not in issue_images[series_name] or is_cover_a:
                issue_images[series_name][vol] = img_url
        ewb.close()

    # --- Build output ---
    series_list = []
    for name in sorted(all_volumes.keys()):
        vols = sorted(all_volumes[name])
        max_vol = summary_max.get(name, max(vols))
        owned_a = sorted(cover_a.get(name, set()))
        owned_other = sorted(all_volumes[name] - cover_a.get(name, set()))

        series_list.append({
            "id": slugify(name),
            "name": name,
            "publisher": "DC",
            "totalIssues": max_vol,
            "ownedCoverA": owned_a,
            "ownedOther": owned_other,
            "imageUrl": series_images.get(name, ""),
            "issueImages": {str(v): url for v, url in sorted(issue_images.get(name, {}).items())},
        })

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump({"series": series_list}, f, indent=2)

    print(f"Wrote {len(series_list)} series to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
